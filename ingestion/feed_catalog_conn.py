"""
Feed catalog connector — reads the master feed list Excel (inbound or outbound).

Excel structure (per your spec):
  - Sheet 1 (index): one row per feed — Feed Name | Business Function | Domain | Frequency
  - One sheet per feed (sheet name == feed name): the feed's field metadata —
    Field Name | Data Type | Required | Business Meaning | PII flag

Writes:
  - feed_catalog  (one row per feed, with direction inbound|outbound)
  - columns       (the feed's fields, so they appear in lineage + Datapoint 360)

Two files are ingested separately with direction set accordingly:
  INBOUND_FEEDS_XLSX  -> direction=inbound
  OUTBOUND_FEEDS_XLSX -> direction=outbound
"""
from __future__ import annotations
import logging
import os
from pathlib import Path

from openpyxl import load_workbook

from .model import Dataset, Column

log = logging.getLogger("cp.feed_catalog")

DOMAINS = ["positions", "taxlots", "transactions", "cash", "fees", "nav", "gl",
           "accruals", "interest", "dividends", "corporate_actions",
           "settlements", "custody", "performance"]


def _norm(s):
    return str(s).strip() if s is not None else ""


def _header_map(ws):
    """Map normalized header -> column index from the first non-empty row."""
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return {(_norm(h).lower()): i for i, h in enumerate(row) if h is not None}
    return {}


class FeedCatalogConnector:
    def __init__(self, xlsx_path: str, direction: str, resolver,
                 platform_id="SWP", schema="feeds"):
        self.xlsx_path = xlsx_path
        self.direction = direction          # inbound | outbound
        self.resolver = resolver
        self.platform_id = platform_id
        self.schema = schema

    @classmethod
    def from_env(cls, direction: str, resolver):
        key = "INBOUND_FEEDS_XLSX" if direction == "inbound" else "OUTBOUND_FEEDS_XLSX"
        path = os.environ.get(key)
        return cls(path, direction, resolver) if path else None

    def parse(self) -> dict:
        wb = load_workbook(self.xlsx_path, data_only=True, read_only=True)
        sheets = wb.sheetnames
        index_sheet = sheets[0]
        ws = wb[index_sheet]
        hm = _header_map(ws)

        def col(row, *names):
            for n in names:
                if n in hm and hm[n] < len(row):
                    return _norm(row[hm[n]])
            return ""

        feeds, columns = [], []
        # Sheet 1: one row per feed
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            feed_name = col(row, "feed name", "feed", "name")
            if not feed_name:
                continue
            biz_fn = col(row, "business function", "function", "business")
            domain = col(row, "domain") or self._infer_domain(feed_name)
            freq = col(row, "frequency", "schedule", "freq")
            feed_id = feed_name.replace(" ", "_")[:200]
            feeds.append({
                "feed_id": feed_id, "feed_name": feed_name[:400],
                "direction": self.direction,
                "business_domain": (domain or None) and domain[:120],
                "frequency": freq[:60] or None,
                "format": None, "record_type": biz_fn[:200] or None,
                "source_system": "SWP" if self.direction == "inbound" else "CP",
                "target_system": "CP" if self.direction == "inbound" else "SWP",
                "schema_ref": feed_name if feed_name in sheets else None,
                "description": biz_fn or None,
                "project_id": self.resolver.resolve_for_swp_feed(feed_name),
                "source_xlsx": self.xlsx_path,
            })

            # per-feed detail sheet (sheet name == feed name)
            if feed_name in sheets:
                fws = wb[feed_name]
                fhm = _header_map(fws)
                pos = 0
                for frow in fws.iter_rows(min_row=2, values_only=True):
                    if not frow or not any(frow):
                        continue
                    def fc(*names):
                        for n in names:
                            if n in fhm and fhm[n] < len(frow):
                                return _norm(frow[fhm[n]])
                        return ""
                    fname = fc("field name", "field", "column", "name")
                    if not fname:
                        continue
                    pos += 1
                    is_pii = fc("pii flag", "pii", "is_pii")
                    columns.append({
                        "dataset_key": f"{self.platform_id}.{self.schema}.{feed_id}".lower(),
                        "feed_id": feed_id,
                        "_object": feed_id,
                        "name": fname[:256],
                        "data_type": fc("data type", "type", "datatype")[:120] or None,
                        "nullable": "N" if fc("required").lower() in ("y", "yes", "true", "required") else "Y",
                        "business_desc": fc("business meaning", "business", "meaning", "description")[:2000] or None,
                        "is_pii": "Y" if is_pii.lower() in ("y", "yes", "true", "pii") else "N",
                        "position_order": pos,
                    })
        wb.close()
        log.info("feed_catalog(%s): %d feeds, %d fields", self.direction,
                 len(feeds), len(columns))
        return {"feeds": feeds, "columns": columns}

    def _infer_domain(self, feed_name):
        fn = feed_name.lower().replace(" ", "_")
        return next((d for d in DOMAINS if d in fn or d.rstrip("s") in fn), None)

    def load(self, loader, bundle):
        for f in bundle["feeds"]:
            loader._merge("feed_catalog", ("feed_id", "direction"), f,
                          protect=("business_domain",))
        # also register feed fields as columns under a FEED dataset, so they
        # flow into lineage and the Datapoint 360 index.
        # Map connector dict -> the real `columns` schema:
        #   platform_id, schema_name, object_name, column_name, data_type,
        #   nullable, is_pii, business_desc, position_order
        for c in bundle["columns"]:
            c.pop("feed_id", None)
            c.pop("dataset_key", None)
            row = {
                "platform_id": self.platform_id,
                "schema_name": self.schema,
                "object_name": (c.get("_object") or "").upper() or None,
                "column_name": c.get("name"),
                "data_type": c.get("data_type"),
                "nullable": c.get("nullable"),
                "is_pii": c.get("is_pii"),
                "business_desc": c.get("business_desc"),
                "position_order": c.get("position_order"),
            }
            loader._merge("columns",
                          ("platform_id", "schema_name", "object_name", "column_name"),
                          row, protect=("is_pii", "pii_attribute"))
        loader.commit()
