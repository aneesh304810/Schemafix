"""Feed Dictionary connector — parses SWP_EOD_Data_Feeds.xlsx multi-tab.

Each data tab becomes a FEED dataset (project='sei'); rows become columns with
parsed data types, PK detection, and enumeration extraction.
"""
from __future__ import annotations
import os
import re
import logging
from dataclasses import dataclass, field
from typing import Optional

from .base import BaseConnector
from .model import Dataset, Column
from .project_resolver import ProjectResolver

log = logging.getLogger("cp.feed_dictionary")

SUPPLEMENTAL_FEEDS = {
    "End of Period Value Aggregation", "Fee Computation", "Fee Group", "Fee Package",
    "Fee Package Usage", "FSL Tax Data", "Custody And Nostro Positions", "Relationships",
    "Account Optional Fields", "Current and Upcoming Activities Impacting Cash",
    "Performance Data Extract", "Contact Details", "Asset Optional Fields",
    "Asset Investment Classification", "User Detail", "User Team and Role",
    "Role Details", "Model", "Model Allocation", "Statement Package", "Statement Event",
    "Statement Event Instance", "Interest Rates", "MiFID Investment Decision Maker",
    "Model Drift Report", "Taxlot",
}
UK_FEEDS = {"Transaction Regulatory", "MiFID II Cost and Charges",
            "MiFID II Product Governance", "FSL Tax Data",
            "MiFID Investment Decision Maker"}


@dataclass
class Enumeration:
    column_name: str
    enum_value: str
    enum_label: str


@dataclass
class FeedDictionaryBundle:
    datasets: list[Dataset] = field(default_factory=list)
    enumerations: dict = field(default_factory=dict)  # ds_key -> [Enumeration]


def parse_data_type(raw: Optional[str]):
    """Return (base, length, precision, scale, fmt)."""
    if not raw:
        return None, None, None, None, None
    s = str(raw).strip()
    m = re.match(r"(NUMBER)\((\d+),(\d+)\)", s, re.I)
    if m:
        return "NUMBER", None, int(m.group(2)), int(m.group(3)), None
    m = re.match(r"(NUMBER)\((\d+)\)", s, re.I)
    if m:
        return "NUMBER", None, int(m.group(2)), None, None
    m = re.match(r"(VARCHAR2|CHAR)\((\d+)\)", s, re.I)
    if m:
        return m.group(1).upper(), int(m.group(2)), None, None, None
    m = re.match(r"(DATE)\s+(\S+)", s, re.I)
    if m:
        return "DATE", None, None, None, m.group(2)
    return s.upper(), None, None, None, None


def extract_enumerations(desc: Optional[str], column_name: str) -> list[Enumeration]:
    """Parse 'N (Label)' or 'N - Description' enumeration patterns."""
    if not desc:
        return []
    out = []
    for m in re.finditer(r"(\d+)\s*[\(\u2013\-]\s*([A-Za-z][^\n\)]*)\)?", str(desc)):
        out.append(Enumeration(column_name, m.group(1), m.group(2).strip()))
    return out


class FeedDictionaryConnector(BaseConnector):
    name = "feed_dictionary"

    def __init__(self, xlsx_path: str, resolver: ProjectResolver,
                 platform_id: str = "swp_feeds", schema: str = "SEI"):
        self.xlsx_path = xlsx_path
        self.resolver = resolver
        self.platform_id = platform_id
        self.schema = schema

    @classmethod
    def from_env(cls) -> "FeedDictionaryConnector":
        return cls(
            xlsx_path=os.environ["DATA360_FEED_DICTIONARY_PATH"],
            resolver=ProjectResolver.from_env(),
        )

    def parse(self) -> FeedDictionaryBundle:
        from openpyxl import load_workbook
        wb = load_workbook(self.xlsx_path, data_only=True, read_only=True)
        # formulas pass to parse HYPERLINK targets in the Link column
        wb_f = load_workbook(self.xlsx_path, data_only=False, read_only=False)
        bundle = FeedDictionaryBundle()

        # 1) read the Contents/index sheet -> metadata per feed
        #    columns (real file): Sno | Interface | Link | Description | Workstream | Static vs Financial
        index_meta = self._parse_contents(wb, wb_f)   # interface_lower -> {desc, workstream, class, sheet}
        sheets_lower = {s.lower(): s for s in wb.sheetnames}

        for sheet in wb.sheetnames:
            if sheet.lower() in ("contents", "index", "toc"):
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            ds, enums = self._parse_tab(sheet, rows)
            # enrich from Contents: match by sheet name OR by hyperlink target
            meta = index_meta.get(sheet.lower())
            if not meta:
                # maybe an Interface whose hyperlink points to this sheet
                meta = next((m for m in index_meta.values()
                             if m.get("sheet", "").lower() == sheet.lower()), None)
            if meta:
                ds.tech_desc = meta.get("desc") or ds.tech_desc
                ds.tags = meta.get("workstream") or ds.tags
                if meta.get("klass"):
                    ds.feed_class = meta["klass"].lower()
            if ds.columns:
                bundle.datasets.append(ds)
                if enums:
                    bundle.enumerations[ds.key] = enums
        wb_f.close()
        log.info("feed_dictionary: parsed %d feeds (%d enriched from Contents)",
                 len(bundle.datasets), len(index_meta))
        return bundle

    def _parse_contents(self, wb, wb_f):
        """Parse the Contents index sheet into interface_lower -> metadata."""
        out = {}
        name = next((s for s in wb.sheetnames if s.lower() in ("contents", "index", "toc")), None)
        if not name:
            return out
        wsv = wb[name]
        wsf = wb_f[name]
        rows = list(wsv.iter_rows(values_only=True))
        if not rows:
            return out
        # locate header row + columns
        hdr_idx, hm = 0, {}
        for i, r in enumerate(rows[:4]):
            low = [str(c).strip().lower() if c is not None else "" for c in r]
            if "interface" in low:
                hdr_idx = i
                hm = {h: j for j, h in enumerate(low) if h}
                break
        hl_re = re.compile(r"#'?([^'!]+?)'?!", re.I)
        for ri in range(hdr_idx + 1, len(rows)):
            r = rows[ri]
            if not r or not any(r):
                continue
            def g(*names):
                for n in names:
                    if n in hm and hm[n] < len(r) and r[hm[n]] is not None:
                        return str(r[hm[n]]).strip()
                return ""
            interface = g("interface", "feed name", "name")
            if not interface:
                continue
            # hyperlink target sheet from the Link cell's formula
            sheet_tgt = ""
            li = hm.get("link")
            if li is not None:
                cell = wsf.cell(row=hdr_idx + 1 + (ri - hdr_idx), column=li + 1)
                fm = str(cell.value or "")
                m = hl_re.search(fm)
                if m:
                    sheet_tgt = m.group(1).strip()
                elif getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "location", None):
                    m2 = hl_re.search("#" + cell.hyperlink.location)
                    if m2:
                        sheet_tgt = m2.group(1).strip()
            out[interface.lower()] = {
                "interface": interface,
                "desc": g("description") or None,
                "workstream": g("workstream") or None,
                "klass": g("static vs financial", "static vs financial ", "type") or None,
                "sheet": sheet_tgt,
            }
        return out

    def _parse_tab(self, feed_name: str, rows: list):
        feed_class = "supplemental" if feed_name in SUPPLEMENTAL_FEEDS else "standard"
        geography = "UK" if feed_name in UK_FEEDS else "US"
        reg = "MiFID II" if "MiFID" in feed_name else ("FSL" if "FSL" in feed_name else None)
        # infer business domain from the feed name so it links to a pipeline
        _domains = ["positions", "taxlots", "transactions", "cash", "fees", "nav",
                    "gl", "accruals", "interest", "dividends", "corporate_actions",
                    "settlements", "custody", "performance", "model", "account",
                    "client", "portfolio", "asset", "statement"]
        _fn = feed_name.lower().replace(" ", "_")
        feed_domain = next((d for d in _domains if d in _fn or d.rstrip("s") in _fn), None)
        ds = Dataset(
            platform_id=self.platform_id, schema=self.schema,
            object_name=feed_name, object_type="FEED",
            project_id=self.resolver.resolve_for_swp_feed(feed_name),
            layer="bronze", feed_class=feed_class, geography=geography,
            domain=feed_domain,
            regulatory_scope=reg, source_xlsx_path=self.xlsx_path,
        )
        enums: list[Enumeration] = []
        if not rows:
            return ds, enums

        # header-aware column resolution (real file: Position | Field Name |
        # Field Description | Data Type/Format | Max Length | Nullable | Reference)
        hdr_row, col = 0, {}
        for i, r in enumerate(rows[:4]):
            low = [str(c).strip().lower() if c is not None else "" for c in r]
            if "field name" in low or "field" in low:
                hdr_row = i
                col = {h: j for j, h in enumerate(low) if h}
                break

        def idx(*names, default=None):
            for n in names:
                if n in col:
                    return col[n]
            return default

        i_pos = idx("position", default=0)
        i_name = idx("field name", "field", "column", default=1)
        i_desc = idx("field description", "description", "business meaning", default=2)
        i_type = idx("data type/format", "data type/format.1", "data type", "type", "format", default=3)
        i_len = idx("max length", "length", "max len", default=4)
        i_null = idx("nullable", "null", default=5)
        i_ref = idx("reference", "ref", default=6)

        start = hdr_row + 1 if col else 1
        for r in rows[start:]:
            if not r or i_name >= len(r) or r[i_name] is None or not str(r[i_name]).strip():
                continue
            g = lambda i: r[i] if (i is not None and i < len(r) and r[i] is not None) else None
            name = _s(g(i_name))
            if not name:
                continue
            try:
                pos = int(g(i_pos))
            except (TypeError, ValueError):
                pos = None
            desc = _s(g(i_desc))
            base, length, prec, scale, fmt = parse_data_type(g(i_type))
            try:
                max_len = int(g(i_len)) if g(i_len) is not None else None
            except (TypeError, ValueError):
                max_len = None
            # real file uses "Not Null"/"Null" text
            nullv = str(g(i_null)).strip().lower().replace(" ", "") if g(i_null) else ""
            nullable = nullv not in ("notnull", "n", "no", "false")
            is_ref = str(g(i_ref)).strip().lower() in ("y", "yes", "true") if g(i_ref) else False
            is_pk = (pos == 1 and name.upper().endswith("_ID") and not nullable)
            ds.columns.append(Column(
                name=name, position_order=pos, data_type=_s(g(i_type)),
                base_data_type=base, data_format=fmt, max_length=length or max_len,
                precision=prec, scale=scale, nullable=nullable, is_pk=is_pk,
                is_reference=is_ref, tech_desc=desc))
            enums.extend(extract_enumerations(desc, name))
        return ds, enums


def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None
