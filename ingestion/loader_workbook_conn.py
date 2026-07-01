"""
Loader workbook connector — reads the real CP_Catalog_SEI_Loaders.xlsx.

Outbound loaders (BBH -> SWP). Ingests all 10 sheets:
  Grouping                       -> ldr_catalog (template family fields)
  Loader_Catalog                 -> ldr_catalog (main row per loader)
  Format_Structure               -> ldr_format_structure
  Attributes                     -> ldr_attributes (master)
  Adhoc_Income_Attributes        -> ldr_attributes (merged supplement)
  Custody_Transfer_V2_Attributes -> ldr_attributes (merged supplement)
  Validations                    -> ldr_validations
  Errors_Exceptions              -> ldr_exceptions
  CP_Catalog_Mapping             -> ldr_module_map
  CIFS_Canonical_Mapping         -> ldr_canonical_map

Also registers each loader as a Datapoint 360 outbound source: writes a FEED-like
dataset (object_type='LOADER', direction outbound) + its attributes as columns, so
the datapoint indexer tags these fields as Outbound.
"""
from __future__ import annotations
import logging
import os
import re

from openpyxl import load_workbook

log = logging.getLogger("cp.loader_workbook")

# header-name -> our field, tolerant of variants. lower-cased contains-match.
def _hdr_index(ws, max_scan=4):
    for r in range(1, max_scan + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, (ws.max_column or 1) + 1)]
        low = [str(v).strip().lower() if v is not None else "" for v in vals]
        if any(low):
            # heuristic: header row has >=2 non-empty cells and includes a known token
            if sum(1 for x in low if x) >= 2:
                return r, {h: i for i, h in enumerate(low) if h}
    return 1, {}


def _s(v):
    return str(v).strip() if v is not None else ""


def _lid(name):
    return re.sub(r"\s+", "_", _s(name))[:200]


class LoaderWorkbookConnector:
    def __init__(self, xlsx_path: str, resolver=None, project_id="sei"):
        self.xlsx_path = xlsx_path
        self.resolver = resolver
        self.project_id = project_id

    @classmethod
    def from_env(cls, resolver=None):
        path = os.environ.get("LOADER_WORKBOOK_XLSX") or os.environ.get("LOADER_CATALOG_XLSX")
        return cls(path, resolver) if path else None

    def parse(self) -> dict:
        wb = load_workbook(self.xlsx_path, data_only=True, read_only=True)
        sheets = {s.lower(): s for s in wb.sheetnames}

        def sheet(*names):
            for n in names:
                if n.lower() in sheets:
                    return wb[sheets[n.lower()]]
            return None

        def rows_of(ws):
            """Yield dict rows keyed by lowercased header."""
            if ws is None:
                return
            hdr, hm = _hdr_index(ws)
            inv = {i: h for h, i in hm.items()}
            for r in range(hdr + 1, (ws.max_row or hdr) + 1):
                vals = [ws.cell(row=r, column=c).value for c in range(1, (ws.max_column or 1) + 1)]
                if not any(v is not None and _s(v) for v in vals):
                    continue
                yield {inv.get(i, f"col{i}"): _s(v) for i, v in enumerate(vals)}

        def pick(row, *names):
            for n in names:
                for k, v in row.items():
                    if n in k:
                        return v
            return ""

        catalog = {}      # loader_id -> ldr_catalog row
        fmt, attrs, vals, excs, modmap, canon = [], {}, [], [], [], []

        # --- Grouping (template families) ---
        for r in rows_of(sheet("Grouping")):
            ln = pick(r, "loader_name", "loader")
            if not ln:
                continue
            lid = _lid(ln)
            c = catalog.setdefault(lid, {"loader_id": lid, "loader_name": ln, "direction": "outbound",
                                         "project_id": self.project_id, "source_xlsx": self.xlsx_path})
            c["group_name"] = pick(r, "group_name", "group") or c.get("group_name")
            c["template_pattern"] = pick(r, "template_pattern", "template") or c.get("template_pattern")
            c["internal_consistency"] = pick(r, "internal_consistency", "consistency") or c.get("internal_consistency")
            c["notes"] = pick(r, "notes") or c.get("notes")

        # --- Loader_Catalog (main) ---
        for r in rows_of(sheet("Loader_Catalog", "Loader Catalog")):
            ln = pick(r, "loader_name", "loader")
            if not ln:
                continue
            lid = _lid(ln)
            c = catalog.setdefault(lid, {"loader_id": lid, "loader_name": ln, "direction": "outbound",
                                         "project_id": self.project_id, "source_xlsx": self.xlsx_path})
            c.update({
                "purpose": pick(r, "purpose"),
                "business_domain": pick(r, "business_domain", "business_doma", "domain"),
                "file_format": pick(r, "file_format", "file_form", "format"),
                "ui_support": pick(r, "ui_support", "ui_suppor", "ui support"),
                "system_support": pick(r, "system_support", "system_supp", "system support"),
                "function_point": pick(r, "function_point", "function point"),
                "header_req": pick(r, "header_req", "header"),
                "footer_req": pick(r, "footer_req", "footer"),
                "date_format": pick(r, "date_fo", "date_format", "date format"),
                "multi_firm": pick(r, "multi_firm", "multi firm"),
                "approval_req": pick(r, "approval_req", "approval"),
                "version": pick(r, "version", "vers"),
                "group_name": pick(r, "group") or c.get("group_name"),
            })

        # --- Format_Structure ---
        for r in rows_of(sheet("Format_Structure", "Format Structure")):
            ln = pick(r, "loader_name", "loader")
            comp = pick(r, "component")
            if not ln or not comp:
                continue
            fmt.append({"loader_id": _lid(ln), "component": comp[:200],
                        "required_for_ui": pick(r, "required_for_u", "required_for_ui", "ui"),
                        "required_for_system": pick(r, "required_for_sy", "required_for_system", "system"),
                        "notes": pick(r, "notes")[:1000] or None})

        # --- Attributes (master) + per-loader supplement sheets, MERGED ---
        def absorb_attrs(ws, src):
            for r in rows_of(ws):
                ln = pick(r, "loader_name", "loader")
                an = pick(r, "attribute_name", "attribute")
                if not ln or not an:
                    continue
                key = (_lid(ln), an[:256])
                row = {
                    "loader_id": key[0], "attribute_name": key[1],
                    "description": pick(r, "description")[:1000] or None,
                    "data_type": pick(r, "data_type", "data_typ", "data type")[:120] or None,
                    "max_length": pick(r, "max_length", "max_leng", "max len")[:60] or None,
                    "optionality": pick(r, "optionality", "optionalit", "optional")[:60] or None,
                    "valid_values": pick(r, "valid_values", "valid_value", "valid")[:1000] or None,
                    "notes": pick(r, "notes")[:1000] or None,
                    "source_sheet": src,
                }
                # merge: later non-empty values fill blanks
                if key in attrs:
                    for k, v in row.items():
                        if v and not attrs[key].get(k):
                            attrs[key][k] = v
                else:
                    attrs[key] = row
        absorb_attrs(sheet("Attributes"), "Attributes")
        absorb_attrs(sheet("Adhoc_Income_Attributes", "Adhoc Income Attributes"), "Adhoc_Income_Attributes")
        absorb_attrs(sheet("Custody_Transfer_V2_Attributes", "Custody Transfer V2 Attributes"), "Custody_Transfer_V2_Attributes")

        # --- Validations ---
        seqv = {}
        for r in rows_of(sheet("Validations")):
            ln = pick(r, "loader_name", "loader")
            an = pick(r, "attribute")
            rule = pick(r, "validation_rule", "validation")
            if not ln or not rule:
                continue
            lid = _lid(ln)
            seqv[lid] = seqv.get(lid, 0) + 1
            vals.append({"loader_id": lid, "attribute_name": an[:256] or "(loader)",
                         "validation_rule": rule[:1000], "error_message": pick(r, "error_message", "error")[:1000] or None,
                         "seq": seqv[lid]})

        # --- Errors_Exceptions ---
        seqe = {}
        for r in rows_of(sheet("Errors_Exceptions", "Errors Exceptions")):
            ln = pick(r, "loader_name", "loader")
            et = pick(r, "exception_type", "exception")
            if not ln or not et:
                continue
            lid = _lid(ln)
            seqe[lid] = seqe.get(lid, 0) + 1
            excs.append({"loader_id": lid, "exception_type": et[:200],
                         "description": pick(r, "description")[:1000] or None,
                         "resolution_path": pick(r, "resolution_path", "resolution")[:1000] or None,
                         "seq": seqe[lid]})

        # --- CP_Catalog_Mapping ---
        for r in rows_of(sheet("CP_Catalog_Mapping", "CP Catalog Mapping")):
            ln = pick(r, "loader_name", "loader")
            if not ln:
                continue
            modmap.append({"loader_id": _lid(ln),
                           "system_interface_360": pick(r, "system_interface", "interface_36", "interface")[:200] or None,
                           "api_360": pick(r, "api_360", "api 360", "api")[:400] or None,
                           "data_360": pick(r, "data_360", "data 360", "data")[:400] or None,
                           "datapoint_360": pick(r, "datapoint_360", "datapoint")[:400] or None,
                           "notes": pick(r, "notes")[:1000] or None})

        # --- CIFS_Canonical_Mapping ---
        seqc = {}
        for r in rows_of(sheet("CIFS_Canonical_Mapping", "CIFS Canonical Mapping")):
            cf = pick(r, "canonical_field", "canonical field")
            ln = pick(r, "loader_name", "loader")
            if not cf or not ln:
                continue
            lid = _lid(ln)
            k = (cf, lid)
            seqc[k] = seqc.get(k, 0) + 1
            canon.append({"canonical_field": cf[:256], "loader_id": lid,
                          "canonical_category": pick(r, "canonical_categ", "category")[:120] or None,
                          "canonical_data_type": pick(r, "canonical_data_typ", "canonical_data")[:120] or None,
                          "physical_field": pick(r, "physical_field", "physical")[:400] or None,
                          "notes": pick(r, "notes")[:1000] or None, "seq": seqc[k]})

        wb.close()
        bundle = {"catalog": list(catalog.values()), "format": fmt,
                  "attributes": list(attrs.values()), "validations": vals,
                  "exceptions": excs, "module_map": modmap, "canonical": canon}
        log.info("loader workbook: %d loaders, %d attrs, %d validations, %d exceptions, "
                 "%d module-maps, %d canonical", len(bundle["catalog"]), len(bundle["attributes"]),
                 len(vals), len(excs), len(modmap), len(canon))
        return bundle

    def load(self, loader, bundle):
        for c in bundle["catalog"]:
            loader._merge("ldr_catalog", ("loader_id",), c)
        for f in bundle["format"]:
            loader._merge("ldr_format_structure", ("loader_id", "component"), f)
        for a in bundle["attributes"]:
            loader._merge("ldr_attributes", ("loader_id", "attribute_name"), a)
        for v in bundle["validations"]:
            loader._merge("ldr_validations", ("loader_id", "attribute_name", "seq"), v)
        for e in bundle["exceptions"]:
            loader._merge("ldr_exceptions", ("loader_id", "seq"), e)
        for m in bundle["module_map"]:
            loader._merge("ldr_module_map", ("loader_id",), m)
        for cm in bundle["canonical"]:
            loader._merge("ldr_canonical_map", ("canonical_field", "loader_id", "seq"), cm)

        # register each loader's attributes as OUTBOUND columns for Datapoint 360,
        # and mirror into feed_catalog(direction=outbound) so direction tagging works.
        for c in bundle["catalog"]:
            loader._merge("feed_catalog", ("feed_id", "direction"), {
                "feed_id": c["loader_id"], "feed_name": c["loader_name"],
                "direction": "outbound", "business_domain": c.get("business_domain"),
                "frequency": None, "format": c.get("file_format"),
                "record_type": c.get("purpose"), "source_system": "BBH",
                "target_system": "SWP", "schema_ref": None,
                "description": c.get("purpose"), "project_id": self.project_id,
                "source_xlsx": self.xlsx_path,
            })
        for a in bundle["attributes"]:
            loader._merge("columns",
                          ("platform_id", "schema_name", "object_name", "column_name"), {
                "platform_id": "SWP",
                "schema_name": "LOADERS",
                "object_name": (a["loader_id"] or "").upper(),
                "column_name": a["attribute_name"],
                "data_type": a.get("data_type"),
                "business_desc": a.get("description"),
                "nullable": "N" if (a.get("optionality") or "").lower().startswith("mandatory") else "Y",
            })
        loader.commit()
